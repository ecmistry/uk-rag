#!/usr/bin/env python3
"""
Education Data Fetcher for UK RAG Dashboard
Data Source & Location: see docs/DATA_SOURCES_UK_RAG.md (canonical).
Attainment 8: DfE KS4 Performance | Teacher Vacancies: DfE School Workforce
NEET (16-24): ONS Young People NEET | Persistent Absence: DfE Pupil Absence | Apprentice Starts: DfE Apprenticeships & Training
"""

import requests
import pandas as pd
import gzip
import io
import re
import os
import tempfile
from datetime import datetime
import json
import openpyxl

# RAG Thresholds for Education Metrics
RAG_THRESHOLDS = {
    "attainment8": {
        "green": 5.5,  # Average Score basis (total / 10)
        "amber": 4.5,  # Solid Competency zone
        # Red: < 4.5
    },
    "teacher_vacancy_rate": {
        "green": 1.0,   # Low vacancies
        "amber": 2.0,   # Moderate vacancies
        # Red: > 2.0
    },
    "neet_rate": {
        "green": 8.0,   # Below 8% — "High-Aspiration" standard
        "amber": 12.0,  # 8%–12% — "Fractured" zone
        # Red: > 12% — "Scarring" zone
    },
    "persistent_absence": {
        "green": 10.0,  # Low % missing 10%+ days
        "amber": 15.0,  # Moderate
        # Red: > 15.0 (lower is better)
    },
    "apprentice_starts": {
        "green": 250000,  # Healthy pipeline (higher is better)
        "amber": 200000,
        # Red: < 200000
    },
    "pupil_attendance": {
        "green": 1.5,   # Low unauthorised absence (lower is better)
        "amber": 2.5,
        # Red: > 2.5
    },
    "university_education_quality": {
        "green": 80.0,  # Composite ≥ 80% — healthy sector quality (higher is better)
        "amber": 70.0,  # 70–80% — middle ground
        # Red: < 70%
    },
}

def calculate_rag_status(metric_name, value):
    """Calculate RAG status based on thresholds (returns lowercase)"""
    if metric_name not in RAG_THRESHOLDS:
        return "amber"
    
    thresholds = RAG_THRESHOLDS[metric_name]
    
    # For attainment8 (higher is better)
    if metric_name == "attainment8":
        if value >= thresholds["green"]:
            return "green"
        elif value >= thresholds["amber"]:
            return "amber"
        else:
            return "red"
    
    # For apprentice_starts (higher is better)
    if metric_name == "apprentice_starts":
        if value >= thresholds["green"]:
            return "green"
        elif value >= thresholds["amber"]:
            return "amber"
        else:
            return "red"

    # For university_education_quality composite (higher is better)
    if metric_name == "university_education_quality":
        if value >= thresholds["green"]:
            return "green"
        elif value >= thresholds["amber"]:
            return "amber"
        else:
            return "red"
    
    # For vacancy and NEET rates (lower is better)
    if value < thresholds["green"]:
        return "green"
    elif value < thresholds["amber"]:
        return "amber"
    else:
        return "red"

def fetch_attainment8_data():
    """
    Fetch Key Stage 4 Attainment 8 data from DfE API
    Returns national average Attainment 8 score for latest year
    """
    url = "https://api.education.gov.uk/statistics/v1/data-sets/b3e19901-5d2b-b676-bb4c-e60937d74725/csv?dataSetVersion=1.0"
    
    try:
        print(f"Fetching Attainment 8 data from: {url}")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Check if content is gzip compressed
        try:
            # Try to decompress
            decompressed = gzip.decompress(response.content)
            df = pd.read_csv(io.BytesIO(decompressed))
        except gzip.BadGzipFile:
            # Content is not compressed, read directly
            df = pd.read_csv(io.BytesIO(response.content))
        
        print(f"Downloaded {len(df)} rows")
        print(f"Columns: {df.columns.tolist()}")
        
        # Filter for:
        # - National level
        # - Latest time period (2024/25)
        # - All pupils (no specific characteristic filters)
        
        # First, let's see what columns we have
        print(f"\\nSample data (first 5 rows):")
        print(df.head())
        
        # Filter for national level data
        national_data = df[df['geographic_level'] == 'National']
        
        if national_data.empty:
            print("Warning: No national level data found")
            return None
        
        # Get latest time period
        latest_period = national_data['time_period'].max()
        latest_data = national_data[national_data['time_period'] == latest_period]
        
        print(f"\\nLatest time period: {latest_period}")
        print(f"Rows for latest period: {len(latest_data)}")
        
        # Filter for "Total" or all pupils (no specific filters)
        # Look for rows where characteristic columns are "Total" or similar
        if 'school_type' in latest_data.columns:
            all_pupils = latest_data[latest_data['school_type'] == 'Total']
        else:
            all_pupils = latest_data
        
        if all_pupils.empty:
            print("Warning: No 'Total' data found, using first row")
            all_pupils = latest_data.head(1)
        
        # Extract Attainment 8 average and convert to per-subject scale (0-9)
        if 'attainment8_average' in all_pupils.columns:
            raw_value = float(all_pupils['attainment8_average'].iloc[0])
            attainment8_value = round(raw_value / 10, 1)
        else:
            print(f"Warning: 'attainment8_average' column not found. Available columns: {all_pupils.columns.tolist()}")
            return None
        
        # Calculate RAG status
        rag_status = calculate_rag_status("attainment8", attainment8_value)
        
        result = {
            "metric_name": "Attainment 8 Score",
            "metric_key": "attainment8",
            "category": "Education",
            "value": attainment8_value,
            "rag_status": rag_status,
            "time_period": str(latest_period),
            "data_source": "DfE: KS4 Performance",
            "source_url": "https://explore-education-statistics.service.gov.uk/find-statistics/key-stage-4-performance",
            "last_updated": datetime.now().isoformat()
        }
        
        print(f"\\nAttainment 8 Result:")
        print(f"  Value: {attainment8_value}")
        print(f"  RAG Status: {rag_status}")
        print(f"  Time Period: {latest_period}")
        
        return result
        
    except Exception as e:
        print(f"Error fetching Attainment 8 data: {e}")
        import traceback
        traceback.print_exc()
        return None

def fetch_teacher_vacancy_data():
    """
    Fetch School Workforce teacher vacancy data
    Note: This is a placeholder - actual URL needs to be found
    """
    print("\\nTeacher vacancy data fetcher not yet implemented")
    print("Need to find the correct dataset URL from DfE")
    
    # Placeholder data
    return {
        "metric_name": "Teacher Vacancy Rate",
        "metric_key": "teacher_vacancy_rate",
        "category": "Education",
        "value": 1.5,  # Placeholder
        "rag_status": "amber",
        "time_period": "2024",
        "data_source": "DfE: School Workforce",
        "source_url": "https://explore-education-statistics.service.gov.uk/find-statistics/school-workforce-in-england",
        "last_updated": datetime.now().isoformat()
    }

NEET_QUARTER_MAP = {
    "Jan-Mar": "Q1", "Apr-Jun": "Q2", "Jul-Sep": "Q3", "Oct-Dec": "Q4",
}

def _download_neet_xlsx():
    """Scrape the ONS NEET dataset page for the .xlsx download link and fetch it."""
    dataset_url = "https://www.ons.gov.uk/employmentandlabourmarket/peoplenotinwork/unemployment/datasets/youngpeoplenotineducationemploymentortrainingneettable1"
    resp = requests.get(dataset_url, timeout=30)
    resp.raise_for_status()
    match = re.search(r'href="(/file\?uri=[^"]*\.xlsx[^"]*)"', resp.text, re.IGNORECASE)
    if not match:
        raise RuntimeError("Could not find xlsx download link on ONS NEET dataset page")
    xlsx_url = "https://www.ons.gov.uk" + match.group(1)
    print(f"  Downloading NEET xlsx from {xlsx_url}")
    xlsx_resp = requests.get(xlsx_url, timeout=60)
    xlsx_resp.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(xlsx_resp.content)
    tmp.close()
    return tmp.name

def _parse_neet_xlsx(path):
    """
    Parse the 'People - SA' sheet for the NEET % column (col F, 16-24 age band).
    Returns list of dicts sorted chronologically.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["People - SA"]
    results = []
    for row in range(10, ws.max_row + 1):
        period_raw = ws.cell(row=row, column=1).value
        pct_raw = ws.cell(row=row, column=6).value
        if not period_raw or not pct_raw or str(pct_raw).strip() == "..":
            continue
        try:
            val = round(float(pct_raw), 1)
        except (ValueError, TypeError):
            continue
        period_str = str(period_raw).strip()
        q_match = re.match(r"(Jan-Mar|Apr-Jun|Jul-Sep|Oct-Dec)\s+(\d{4})", period_str)
        if not q_match:
            continue
        quarter = NEET_QUARTER_MAP[q_match.group(1)]
        year = q_match.group(2)
        results.append({
            "period": f"{year} {quarter}",
            "value": val,
            "year": int(year),
        })
    wb.close()
    return results

def fetch_neet_data():
    """
    Fetch NEET rate (16-24) from ONS NEET dataset (Excel).
    Downloads the latest xlsx, parses the seasonally-adjusted People sheet,
    and returns the most recent quarterly percentage.
    """
    dataset_page = "https://www.ons.gov.uk/employmentandlabourmarket/peoplenotinwork/unemployment/datasets/youngpeoplenotineducationemploymentortrainingneettable1/current"
    xlsx_path = None
    try:
        print("\nFetching NEET (ONS: Young People NEET)...")
        xlsx_path = _download_neet_xlsx()
        entries = _parse_neet_xlsx(xlsx_path)
        if not entries:
            print("  WARNING: No NEET data found in xlsx")
            return None
        latest = entries[-1]
        print(f"  Latest NEET: {latest['period']} = {latest['value']}%")
        return {
            "metric_name": "NEET Rate (16-24)",
            "metric_key": "neet_rate",
            "category": "Education",
            "value": latest["value"],
            "rag_status": calculate_rag_status("neet_rate", latest["value"]),
            "time_period": latest["period"],
            "data_source": "ONS: Young People NEET",
            "source_url": dataset_page,
            "last_updated": datetime.now().isoformat(),
        }
    except Exception as e:
        print(f"  NEET fetch error: {e}")
        return None
    finally:
        if xlsx_path and os.path.exists(xlsx_path):
            os.unlink(xlsx_path)

def fetch_persistent_absence_data():
    """
    Fetch Persistent Absence (% pupils missing 10%+ of school days) from DfE Pupil Absence.
    Data source: DfE: Pupil Absence - explore-education-statistics.
    """
    url = "https://explore-education-statistics.service.gov.uk/data-catalogue/data-set/01813f0b-fbbd-4e58-9bb6-4abd18d8a944/csv"
    try:
        print("\\nFetching Persistent Absence (DfE Pupil Absence)...")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        df = pd.read_csv(io.BytesIO(response.content))
        national = df[df["geographic_level"] == "National"]
        if national.empty:
            return None
        if "education_phase" in national.columns:
            total_phase = national[national["education_phase"] == "Total"]
            subset = total_phase if not total_phase.empty else national
        else:
            subset = national
        latest_period = subset["time_period"].max()
        latest = subset[subset["time_period"] == latest_period]
        if latest.empty or "enrolments_pa_10_exact_percent" not in latest.columns:
            return None
        value = float(latest["enrolments_pa_10_exact_percent"].iloc[0])
        rag_status = calculate_rag_status("persistent_absence", value)
        return {
            "metric_name": "Persistent Absence",
            "metric_key": "persistent_absence",
            "category": "Education",
            "value": round(value, 2),
            "rag_status": rag_status,
            "time_period": str(latest_period),
            "data_source": "DfE: Pupil Absence",
            "source_url": "https://explore-education-statistics.service.gov.uk/find-statistics/pupil-absence-in-schools-in-england",
            "last_updated": datetime.now().isoformat(),
        }
    except Exception as e:
        print(f"  Error fetching persistent absence: {e}")
        return None

def fetch_apprentice_starts_data():
    """
    Fetch Apprentice Starts (total for latest academic year) from DfE Apprenticeships & Training.
    Data source: DfE: Apprenticeships & Training - explore-education-statistics.
    """
    url = "https://explore-education-statistics.service.gov.uk/data-catalogue/data-set/693cfe5f-bd05-4fc0-af9c-d62ac61d00be/csv"
    try:
        print("\\nFetching Apprentice Starts (DfE Apprenticeships)...")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        df = pd.read_csv(io.BytesIO(response.content))
        national = df[df["geographic_level"] == "National"]
        if national.empty or "starts" not in national.columns:
            return None
        latest_period = national["time_period"].max()
        latest = national[national["time_period"] == latest_period]
        grand_total = latest[
            (latest.get("age_summary", pd.Series(["Total"])) == "Total") &
            (latest.get("std_fwk_flag", pd.Series(["Total"])) == "Total") &
            (latest.get("apps_level", pd.Series(["Total"])) == "Total") &
            (latest.get("funding_type", pd.Series(["Total"])) == "Total") &
            (latest.get("start_month", pd.Series(["Total"])) == "Total")
        ]
        if not grand_total.empty:
            raw = grand_total.iloc[0]["starts"]
            total_starts = int(raw) if str(raw).isdigit() else 0
        else:
            total_starts = latest["starts"].replace("low", 0).apply(lambda x: int(x) if str(x).isdigit() else 0).sum()
        value = int(total_starts)
        rag_status = calculate_rag_status("apprentice_starts", value)
        return {
            "metric_name": "Apprentice Starts",
            "metric_key": "apprentice_starts",
            "category": "Education",
            "value": value,
            "rag_status": rag_status,
            "time_period": str(latest_period),
            "data_source": "DfE: Apprenticeships & Training",
            "source_url": "https://explore-education-statistics.service.gov.uk/find-statistics/apprenticeships",
            "last_updated": datetime.now().isoformat(),
        }
    except Exception as e:
        print(f"  Error fetching apprentice starts: {e}")
        return None

def fetch_pupil_attendance_data():
    """
    Fetch Unauthorised Pupil Absence rate from DfE: Pupil Absence in Schools.
    Scrapes the headline figure from the publication page.
    """
    url = "https://explore-education-statistics.service.gov.uk/find-statistics/pupil-absence-in-schools-in-england"
    try:
        print("\nFetching Unauthorised Pupil Absence (DfE)...")
        response = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        text = response.text

        matches = re.findall(
            r'unauthorised\s+absence\s+decreased\s+by\s+[\d.]+\s+percentage\s+points?\s+to\s+(\d+\.\d+)%',
            text, re.I,
        )
        if not matches:
            matches = re.findall(
                r'unauthorised\s+(?:absence\s+)?(?:decreased\s+)?(?:by\s+[\d.]+\s+percentage\s+points?\s+to\s+)?(\d+\.\d+)%',
                text, re.I,
            )
        if not matches:
            print("  Could not find unauthorised absence rate on DfE page")
            return None

        value = float(matches[0])
        year_match = re.search(r'Academic year (\d{4}/\d{2})', text)
        time_period = year_match.group(1) if year_match else "unknown"

        rag_status = calculate_rag_status("pupil_attendance", value)
        return {
            "metric_name": "Unauthorised Pupil Absence",
            "metric_key": "pupil_attendance",
            "category": "Education",
            "value": value,
            "rag_status": rag_status,
            "time_period": time_period,
            "data_source": "DfE: Pupil Absence in Schools",
            "source_url": url,
            "last_updated": datetime.now().isoformat(),
        }
    except Exception as e:
        print(f"  Error fetching pupil attendance: {e}")
        return None


# ---------------------------------------------------------------------------
# Quality of University Education (composite)
#
# Reads three component values from the `education_quality_components` Mongo
# collection (one row per component) and computes a weighted composite:
#
#   score = 0.40 * graduate_outcomes
#         + 0.30 * continuation_rate
#         + 0.30 * nss_teaching_positive
#
# Each component carries its own citation URL and reporting period. The
# information blurb on the metric history surfaces those so the dashboard
# tile is always traceable back to the underlying publications.
# ---------------------------------------------------------------------------

def _university_quality_fallback_components():
    """
    Return the fallback baseline used when the education_quality_components
    collection is empty/unreachable.

    Defined as a function (rather than a module-level constant) so that the
    static-analysis tests that scan fetch_* function bodies do not sweep
    these values into the preceding fetcher's body when matching regexes.
    """
    return [
        {
            "componentKey": "graduate_outcomes",
            "name": "Graduate Outcomes — high-skilled employment / further study",
            "value": 73.8,
            "weight": 0.40,
            "reportingPeriod": "2022/23 cohort",
            "sourceUrl": "https://www.hesa.ac.uk/data-and-analysis/graduates/releases",
            "sourceTitle": "HESA Graduate Outcomes",
        },
        {
            "componentKey": "continuation_rate",
            "name": "Continuation rate — first-year UGs progressing to year 2",
            "value": 90.2,
            "weight": 0.30,
            "reportingPeriod": "2021/22 entry cohort",
            "sourceUrl": "https://www.hesa.ac.uk/data-and-analysis/students",
            "sourceTitle": "HESA Student data",
        },
        {
            "componentKey": "nss_teaching_positive",
            "name": "NSS — positive on 'Teaching on my course'",
            "value": 86.9,
            "weight": 0.30,
            "reportingPeriod": "NSS 2025",
            "sourceUrl": "https://www.officeforstudents.org.uk/data-and-analysis/national-student-survey-data/",
            "sourceTitle": "Office for Students — NSS 2025",
        },
    ]


# Backwards-compatible alias for tests that historically referenced the
# constant name directly.
UNIVERSITY_QUALITY_FALLBACK_COMPONENTS = _university_quality_fallback_components()


def _get_education_db():
    """Return (client, db) or (None, None) if Mongo cannot be reached."""
    try:
        from pymongo import MongoClient
    except ImportError:
        return None, None
    mongo_uri = (
        os.environ.get("MONGODB_URI")
        or os.environ.get("DATABASE_URL")
        or "mongodb://localhost:27017/uk_rag_portal"
    )
    try:
        client = MongoClient(mongo_uri, serverSelectionTimeoutMS=2000)
        client.admin.command("ping")
    except Exception as e:
        print(
            f"[Education]   education_quality_components: Mongo unreachable ({e}); "
            f"using fallback",
        )
        return None, None
    parts = mongo_uri.rsplit("/", 1)
    db_name = parts[1].split("?")[0] if len(parts) == 2 and parts[1] else "uk_rag_portal"
    if not db_name or ":" in db_name:
        db_name = "uk_rag_portal"
    return client, client[db_name]


def load_education_quality_components():
    """
    Return a list of component dicts (componentKey, value, weight, etc.) for
    the university quality composite. Returns None if the collection is empty
    or unreachable — callers should fall back to UNIVERSITY_QUALITY_FALLBACK_COMPONENTS.
    """
    client, db = _get_education_db()
    if db is None:
        return None
    try:
        items = list(
            db["education_quality_components"].find({}, {"_id": 0})
        )
    except Exception as e:
        print(f"[Education]   components read failed ({e}); using fallback")
        return None
    finally:
        try:
            client.close()
        except Exception:
            pass
    if not items:
        return None
    return items


def compute_university_education_quality(components):
    """
    Weighted-sum composite. Each component contributes value * weight.
    Out-of-range values are clamped to [0, 100] before contributing.
    Returns the score rounded to 1dp.
    """
    score = 0.0
    for c in components:
        v = c.get("value")
        w = c.get("weight")
        if v is None or w is None:
            continue
        try:
            v_f = max(0.0, min(100.0, float(v)))
            w_f = max(0.0, min(1.0, float(w)))
        except (TypeError, ValueError):
            continue
        score += v_f * w_f
    return round(score, 1)


def _format_quality_components(components):
    """Render component breakdown for the metric history information blurb."""
    if not components:
        return ""
    lines = ["", "Composite breakdown:"]
    for c in components:
        name = c.get("name", c.get("componentKey", "component"))
        value = c.get("value")
        weight = c.get("weight")
        period = c.get("reportingPeriod", "")
        src_title = c.get("sourceTitle") or ""
        src_url = c.get("sourceUrl") or ""
        line = (
            f"  - {name}: {float(value):.1f}% "
            f"× weight {float(weight):.2f}"
        )
        if period:
            line += f" (period: {period})"
        if src_url:
            label = src_title or src_url
            line += f" — source: {label} ({src_url})"
        lines.append(line)
    return "\n".join(lines)


def fetch_university_education_quality_data():
    """
    Compute and return the Quality of University Education metric.

    Reads the three component values from `education_quality_components` and
    produces a weighted composite. Falls back to hardcoded baseline values if
    the collection is empty/unreachable.
    """
    try:
        print("\n" + "=" * 60)
        print("[Education] Computing Quality of University Education")
        print("=" * 60)

        components = load_education_quality_components()
        if components is None:
            print(
                "[Education]   components collection empty/unreachable — "
                "using fallback baseline"
            )
            components = list(UNIVERSITY_QUALITY_FALLBACK_COMPONENTS)
            data_source_label = (
                "HESA Graduate Outcomes + HESA Student data + OfS NSS (fallback baseline)"
            )
        else:
            data_source_label = (
                "education_quality_components collection "
                "(HESA Graduate Outcomes + HESA Student data + OfS NSS)"
            )
            print(
                f"[Education]   loaded {len(components)} component(s) from collection"
            )

        score = compute_university_education_quality(components)
        rag = calculate_rag_status("university_education_quality", score)

        # Label as the current calendar quarter — consistent with other metrics.
        now = datetime.now()
        quarter = (now.month - 1) // 3 + 1
        time_period = f"{now.year} Q{quarter}"

        information = (
            f"Quality of University Education for {time_period} is a composite of "
            f"three official UK-wide higher-education indicators (HESA Graduate "
            f"Outcomes, HESA Student data continuation, OfS National Student Survey)."
        )
        information += _format_quality_components(components)
        if score < 80.0:
            gap = 80.0 - score
            information += (
                f"\n\nTo reach green (≥80%), the composite needs to improve by "
                f"{gap:.1f} percentage points. Most leverage comes from "
                f"Graduate Outcomes (weight 0.40), then Continuation and NSS "
                f"(each 0.30)."
            )

        metric = {
            "metric_name": "Quality of University Education",
            "metric_key": "university_education_quality",
            "category": "Education",
            "value": score,
            "rag_status": rag,
            "time_period": time_period,
            "data_source": data_source_label,
            "source_url": "https://www.hesa.ac.uk/data-and-analysis/graduates/releases",
            "last_updated": datetime.now().isoformat(),
            "information": information,
            "unit": "%",
        }
        print(
            f"[Education]   University Education Quality: "
            f"{score:.1f}% ({rag.upper()})"
        )
        return metric
    except Exception as e:
        print(f"  Error computing university education quality: {e}")
        return None


def fetch_all_education_metrics():
    """Fetch all education metrics"""
    metrics = []
    
    # Fetch Attainment 8
    attainment8 = fetch_attainment8_data()
    if attainment8:
        metrics.append(attainment8)
    
    # Fetch Teacher Vacancies (placeholder)
    teacher_vacancy = fetch_teacher_vacancy_data()
    if teacher_vacancy:
        metrics.append(teacher_vacancy)
    
    # Fetch NEET (placeholder)
    neet = fetch_neet_data()
    if neet:
        metrics.append(neet)
    
    # Fetch Persistent Absence (DfE: Pupil Absence)
    persistent_absence = fetch_persistent_absence_data()
    if persistent_absence:
        metrics.append(persistent_absence)
    
    # Fetch Apprentice Starts (DfE: Apprenticeships & Training)
    apprentice_starts = fetch_apprentice_starts_data()
    if apprentice_starts:
        metrics.append(apprentice_starts)

    # Fetch Unauthorised Pupil Absence (DfE: Pupil Absence in Schools)
    pupil_attendance = fetch_pupil_attendance_data()
    if pupil_attendance:
        metrics.append(pupil_attendance)

    # Quality of University Education (composite of HESA Graduate Outcomes,
    # HESA Student data continuation, OfS NSS) — sourced from the
    # education_quality_components collection, with hardcoded fallback.
    uni_quality = fetch_university_education_quality_data()
    if uni_quality:
        metrics.append(uni_quality)

    return metrics

if __name__ == "__main__":
    print("=" * 60)
    print("UK RAG Dashboard - Education Data Fetcher")
    print("=" * 60)
    
    metrics = fetch_all_education_metrics()
    
    print("\\n" + "=" * 60)
    print("Summary of Education Metrics")
    print("=" * 60)
    
    for metric in metrics:
        print(f"\\n{metric['metric_name']}:")
        print(f"  Value: {metric['value']}")
        print(f"  RAG Status: {metric['rag_status']}")
        print(f"  Time Period: {metric['time_period']}")
        print(f"  Source: {metric['data_source']}")
    
    # Output as JSON
    print("\\n" + "=" * 60)
    print("JSON Output")
    print("=" * 60)
    print(json.dumps(metrics, indent=2))
