#!/usr/bin/env python3
"""
Fetch UK public sector expenditure on services from HM Treasury PSS Excel.

Source: https://www.gov.uk/government/collections/national-statistics-release
File:   "Expenditure on Services" (PSS_*_TES.xlsx), sheet Table_10a

Usage:
  python3 server/public_expenditure_fetcher.py --chart   # JSON to stdout for tRPC
  python3 server/public_expenditure_fetcher.py --cron    # Upsert into MongoDB

Crontab (daily at 07:00 UTC):
  0 7 * * * cd /home/ec2-user/uk-rag-portal && python3 server/public_expenditure_fetcher.py --cron >> logs/public_expenditure.log 2>&1
"""

import json
import os
import re
import sys
import tempfile
from datetime import datetime, timezone

import requests

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

COLLECTION_URL = "https://www.gov.uk/government/collections/national-statistics-release"
BASE_URL = "https://www.gov.uk"
USER_AGENT = "UK-RAG-Dashboard/1.0 (public expenditure fetcher)"

# Rows in Table_10a to extract (1-indexed).
# Yellow cells: the 16 expenditure categories.
# Blue cell (row 10): public sector debt interest — stored separately for toggle.
ROWS = {
    8:  {"key": "public_and_common_services",      "label": "General Public Services: Public & Common"},
    9:  {"key": "international_services",           "label": "General Public Services: International"},
    10: {"key": "debt_interest",                    "label": "Public Sector Debt Interest"},
    11: {"key": "defence",                          "label": "Defence"},
    12: {"key": "public_order_and_safety",          "label": "Public Order and Safety"},
    14: {"key": "enterprise_and_economic_dev",      "label": "Economic Affairs: Enterprise & Development"},
    15: {"key": "science_and_technology",           "label": "Economic Affairs: Science & Technology"},
    16: {"key": "employment_policies",              "label": "Economic Affairs: Employment Policies"},
    17: {"key": "agriculture_fisheries_forestry",   "label": "Economic Affairs: Agriculture"},
    18: {"key": "transport",                        "label": "Economic Affairs: Transport"},
    19: {"key": "environment_protection",           "label": "Environment Protection"},
    20: {"key": "housing_and_community",            "label": "Housing and Community Amenities"},
    21: {"key": "health",                           "label": "Health"},
    22: {"key": "recreation_culture_religion",      "label": "Recreation, Culture and Religion"},
    23: {"key": "education",                        "label": "Education"},
    24: {"key": "social_protection",                "label": "Social Protection"},
    25: {"key": "eu_transactions",                  "label": "EU Transactions"},
}

ALL_KEYS = [v["key"] for v in ROWS.values()]


def log(msg):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    print(f"{ts} [ExpenditureFetcher] {msg}", file=sys.stderr, flush=True)


def find_latest_release_url():
    """Scrape the HMT collection page to find the latest PSS release."""
    log("Checking HMT collection page for latest release...")
    resp = requests.get(COLLECTION_URL, headers={"User-Agent": USER_AGENT}, timeout=30)
    resp.raise_for_status()

    links = re.findall(
        r'href="(/government/statistics/public-spending-statistics-release-[^"]+)"',
        resp.text,
    )
    if not links:
        raise RuntimeError("No PSS release links found on collection page")

    month_order = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12,
    }

    def sort_key(path):
        m = re.search(r"release-(\w+)-(\d{4})", path)
        if m:
            return (int(m.group(2)), month_order.get(m.group(1), 0))
        return (0, 0)

    latest = max(set(links), key=sort_key)
    url = f"{BASE_URL}{latest}"
    log(f"Latest release page: {url}")
    return url


def find_tes_xlsx_url(release_url):
    """From a release page, find the Expenditure on Services XLSX download link."""
    resp = requests.get(release_url, headers={"User-Agent": USER_AGENT}, timeout=30)
    resp.raise_for_status()

    matches = re.findall(
        r'(https://assets\.publishing\.service\.gov\.uk/media/[^"]+TES[^"]*\.xlsx)',
        resp.text,
    )
    if not matches:
        all_xlsx = re.findall(
            r'(https://assets\.publishing\.service\.gov\.uk/media/[^"]+\.xlsx)',
            resp.text,
        )
        for url in all_xlsx:
            if "TES" in url.upper():
                return url
        raise RuntimeError(f"No TES XLSX found on {release_url}. XLSX links: {all_xlsx}")

    return matches[0]


def download_excel(url=None):
    """Download the PSS TES Excel file to a temp path."""
    if url is None:
        release_url = find_latest_release_url()
        url = find_tes_xlsx_url(release_url)

    log(f"Downloading {url} ...")
    resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=120)
    resp.raise_for_status()
    if len(resp.content) < 10_000:
        raise ValueError(f"Downloaded file too small ({len(resp.content)} bytes)")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(resp.content)
    tmp.close()
    log(f"Downloaded {len(resp.content):,} bytes to {tmp.name}")
    return tmp.name


def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def extract_expenditure_data(excel_path):
    """
    Read Table_10 (nominal/cash terms), extract the 17 category rows for all
    fiscal years.  We use nominal rather than Table_10a (real terms) because
    the ONS receipts data we compare against is also in nominal terms.
    Returns list of period dicts sorted chronologically.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Table_10"]

    header_row = list(ws.iter_rows(min_row=5, max_row=5))[0]
    fiscal_years = {}
    for cell in header_row[1:]:
        if cell.value and re.match(r"\d{4}-\d{2}", str(cell.value)):
            fiscal_years[cell.column] = str(cell.value)

    if not fiscal_years:
        wb.close()
        raise RuntimeError("No fiscal year headers found in Table_10a row 5")

    log(f"Found {len(fiscal_years)} fiscal years: {list(fiscal_years.values())[0]} to {list(fiscal_years.values())[-1]}")

    data_rows = {}
    for row_num, meta in ROWS.items():
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]
        data_rows[meta["key"]] = row_cells

    periods = []
    for col_idx, fy in sorted(fiscal_years.items()):
        entry = {"period": fy}
        for row_num, meta in ROWS.items():
            row_cells = data_rows[meta["key"]]
            cell_val = row_cells[col_idx - 1].value if col_idx - 1 < len(row_cells) else None
            entry[meta["key"]] = round(safe_float(cell_val), 1)
        periods.append(entry)

    wb.close()
    log(f"Extracted {len(periods)} fiscal years from {periods[0]['period']} to {periods[-1]['period']}")
    return periods


def run_chart_mode(excel_path):
    """Output JSON to stdout for tRPC consumption."""
    data = extract_expenditure_data(excel_path)
    print(json.dumps({"periods": data}))


def run_cron_mode(excel_path):
    """Upsert fiscal year data into MongoDB."""
    try:
        import pymongo
    except ImportError:
        log("ERROR: pymongo is required for --cron mode")
        sys.exit(1)

    mongo_uri = (
        os.environ.get("MONGODB_URI")
        or os.environ.get("DATABASE_URL")
        or "mongodb://localhost:27017/uk_rag_portal"
    )
    client = pymongo.MongoClient(mongo_uri)
    db_name = mongo_uri.rsplit("/", 1)[-1].split("?")[0] if "/" in mongo_uri else "uk_rag_portal"
    db = client[db_name]
    collection = db["publicSectorExpenditure"]

    collection.create_index("period", unique=True)

    data = extract_expenditure_data(excel_path)
    now = datetime.now(timezone.utc)

    upserted = 0
    for entry in data:
        result = collection.update_one(
            {"period": entry["period"]},
            {
                "$set": {**entry, "updatedAt": now},
                "$setOnInsert": {"createdAt": now},
            },
            upsert=True,
        )
        if result.upserted_id or result.modified_count:
            upserted += 1

    log(f"Upserted {upserted} fiscal years into publicSectorExpenditure collection")
    client.close()


def find_local_file():
    """Look for a local PSS TES xlsx in the project root."""
    project_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    for name in sorted(os.listdir(project_root), reverse=True):
        if name.startswith("PSS_") and name.endswith(".xlsx") and "TES" in name.upper():
            path = os.path.join(project_root, name)
            log(f"Using local file: {path}")
            return path
    return None


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ("--chart", "--cron"):
        print("Usage: python3 public_expenditure_fetcher.py [--chart|--cron]", file=sys.stderr)
        sys.exit(1)

    mode = sys.argv[1]

    if mode == "--cron":
        excel_path = download_excel()
        should_cleanup = True
    else:
        local = find_local_file()
        if local:
            excel_path = local
            should_cleanup = False
        else:
            excel_path = download_excel()
            should_cleanup = True

    try:
        if mode == "--chart":
            run_chart_mode(excel_path)
        elif mode == "--cron":
            run_cron_mode(excel_path)
    finally:
        if should_cleanup:
            os.unlink(excel_path)


if __name__ == "__main__":
    main()
